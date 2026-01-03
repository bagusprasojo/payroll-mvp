from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from typing import Iterable

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    Employee,
    PayrollComponent,
    PayrollEntry,
    PayrollEntryItem,
    PayrollPeriod,
    School,
    User,
)


class PayrollGenerationError(Exception):
    """High level error during payroll generation."""


def _ensure_entry(period: PayrollPeriod, employee: Employee) -> PayrollEntry:
    entry, _ = PayrollEntry.objects.get_or_create(period=period, employee=employee)
    entry.status = PayrollEntry.STATUS_DRAFT
    entry.save(update_fields=["status", "updated_at"])
    entry.items.all().delete()
    return entry


def _create_items(entry: PayrollEntry, components: Iterable[PayrollComponent], amount_map: dict[str, Decimal]) -> None:
    items = []
    for component in components:
        amount = amount_map.get(component.code, component.default_amount if component.is_fixed else Decimal("0"))
        items.append(
            PayrollEntryItem(
                entry=entry,
                component=component,
                component_name=component.name,
                component_type=component.component_type,
                amount=amount,
            )
        )
    PayrollEntryItem.objects.bulk_create(items)
    entry.recalculate_totals()


def _copy_from_period(target_period: PayrollPeriod, source_period: PayrollPeriod) -> None:
    for entry in source_period.entries.select_related("employee"):
        target_entry = _ensure_entry(target_period, entry.employee)
        items = [
            PayrollEntryItem(
                entry=target_entry,
                component=item.component,
                component_name=item.component_name,
                component_type=item.component_type,
                amount=item.amount,
            )
            for item in entry.items.all()
        ]
        PayrollEntryItem.objects.bulk_create(items)
        target_entry.recalculate_totals()


def _manual_generation(period: PayrollPeriod, school: School) -> None:
    components = list(school.components.filter(is_active=True))
    employees = list(school.employees.filter(is_active=True))
    for employee in employees:
        entry = _ensure_entry(period, employee)
        _create_items(entry, components, {})


def _import_amounts(upload_file, school: School) -> dict[str, dict[str, Decimal]]:
    try:
        workbook = load_workbook(upload_file, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl specific errors
        raise PayrollGenerationError("File Excel tidak valid.") from exc

    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        email_idx = header.index("email")
        component_idx = header.index("component_code")
        amount_idx = header.index("amount")
    except ValueError as exc:
        raise PayrollGenerationError("Kolom wajib: email, component_code, amount.") from exc

    data: dict[str, dict[str, Decimal]] = defaultdict(dict)
    for row in sheet.iter_rows(min_row=2):
        email = row[email_idx].value
        component_code = row[component_idx].value
        amount = row[amount_idx].value
        if not email or not component_code:
            continue
        try:
            decimal_amount = Decimal(str(amount or 0))
        except Exception as exc:  # pragma: no cover - Decimal conversion
            raise PayrollGenerationError(f"Nominal tidak valid untuk {email}/{component_code}.") from exc
        data[str(email).lower()][str(component_code).upper()] = decimal_amount

    if not data:
        raise PayrollGenerationError("Tidak ada data dalam file.")

    existing_emails = set(
        school.employees.filter(email__in=data.keys()).values_list("email", flat=True)
    )
    missing = set(data.keys()) - existing_emails
    if missing:
        raise PayrollGenerationError(f"Pegawai tidak ditemukan: {', '.join(missing)}")
    return data


def generate_payroll(
    *,
    period: PayrollPeriod,
    method: str,
    school: School,
    user: User | None,
    source_period: PayrollPeriod | None = None,
    upload_file=None,
) -> None:
    components = list(school.components.filter(is_active=True))
    if not components:
        raise PayrollGenerationError("Belum ada komponen gaji aktif.")
    employees = list(school.employees.filter(is_active=True))
    if not employees:
        raise PayrollGenerationError("Belum ada pegawai aktif.")

    with transaction.atomic():
        if method == "manual":
            _manual_generation(period, school)
        elif method == "copy":
            if not source_period:
                raise PayrollGenerationError("Periode sumber wajib diisi untuk copy.")
            _copy_from_period(period, source_period)
        elif method == "import":
            amount_map = _import_amounts(upload_file, school)
            for employee in employees:
                entry = _ensure_entry(period, employee)
                mapped_amounts = amount_map.get(employee.email.lower(), {})
                _create_items(entry, components, mapped_amounts)
        else:
            raise PayrollGenerationError("Metode generate tidak dikenal.")

        period.mark_generated(user)


def add_employee_payroll_entry(*, period: PayrollPeriod, employee: Employee, school: School) -> PayrollEntry:
    if employee.school_id != school.id:
        raise PayrollGenerationError("Pegawai tidak berasal dari sekolah ini.")
    components = list(school.components.filter(is_active=True))
    if not components:
        raise PayrollGenerationError("Belum ada komponen gaji aktif.")
    entry = _ensure_entry(period, employee)
    _create_items(entry, components, {})
    return entry
